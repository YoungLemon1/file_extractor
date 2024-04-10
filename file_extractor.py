# importing required classes 
import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter import messagebox
from tkinter import ttk
from pypdf import PdfReader, PdfWriter 
from openpyxl import load_workbook, utils
from pathlib import Path
import os

### Logic

header_text = "FILENAME"

str_pdf = "documents.pdf"
str_wb = "SQL.xlsx"
contract_pages_count = 5

def confirm_overwrite(outputpdf):
    result = messagebox.askyesnocancel("Confirmation", f"The file {outputpdf} already exists. Do you want to overwrite it?")
    return result

def show_error_message():
    messagebox.showerror("Error: Permission Deined", "Error occured while writing to this file\n(make sure it is not open)")

# Load the workbook
def find_cell_with_header(sheet, header_text):
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=1), start=1):
        for cell in row:
            if cell.value == header_text:
                return {"row": row_number, "column": cell.column_letter}
    return None

def select_pdf_file():
    global str_pdf
    file_path = askopenfilename(parent=root, title="Choose a file", filetype=[("Pdf file", "*.pdf")])
    if file_path:
        source_pdf_text.set(file_path)
        str_pdf = file_path

def select_excel_file():
    global str_wb
    file_path = askopenfilename(parent=root, title="Choose a file", filetype=[("Excel file", "*.xlsx")])
    if file_path:
        file_name = os.path.basename(file_path)
        source_excel_text.set(file_path)
        str_wb = file_path

def pdf_extract():
    def update_extraction_failed(err = "please try again"):
        status_label_text.set(f"Extraction failed: {err}")

    try:
        pdf = PdfReader(str_pdf)
    except:
        update_extraction_failed(f"could not find pdf file {str_pdf}")
        return

    try:
        wb = load_workbook(str_wb, data_only = True)
    except:
        update_extraction_failed(f"could not find excel file {str_wb}")
        return
    # Assuming you're working with the first sheet, change it accordingly if needed
    sheet = wb.active

    contract_pages_count = int(entry_page_per_contract.get())

    pdf_pages_count = len(pdf.pages)

    if pdf_pages_count % contract_pages_count != 0:
        update_extraction_failed("mismatched page count between files")
        return

    p = Path(str_pdf)
    os.chdir(p.parent)
    # output pdf file name
    total_pages = len(pdf.pages)
    contract_count = total_pages // contract_pages_count
    success_count = 0

    header_cell = find_cell_with_header(sheet, header_text)
    if not header_cell:
        update_extraction_failed("could not find FILENAME column in Excel file")
        return
    h_row = header_cell["row"]
    h_column = header_cell["column"]

    row_index = h_row + 1

    if not header_cell:
        print(f"Header cell {header_text} not found")
        return
    else:
        print(f"Header cell {header_text} found: {header_cell}")
    
    for i in range(0, contract_count):
        writer = PdfWriter() 
        # adding pages to pdf writer object
        start_page = i * contract_pages_count
        end_page = start_page + contract_pages_count

        column_index = utils.column_index_from_string(h_column)
        cell_value = sheet.cell(row = row_index, column = column_index).value
        if cell_value:
            outputpdf = f"{cell_value}.pdf"
            if(os.path.exists(outputpdf) and (not check_val.get())):
                confirmation = confirm_overwrite(outputpdf)
                if confirmation == False:
                    continue
                elif confirmation == None:
                    break
        for page in range(start_page, end_page):
            writer.add_page(pdf.pages[page]) 

        # writing split pdf pages to pdf file 
        try:
            with open(outputpdf, "wb") as f:
                writer.write(f)
        except Exception as e:
            # Handle any exceptions that occur during writing
            show_error_message()
            #update_extraction_failed(f"Error occurred while writing {outputpdf}: {str(e)}")
            break
        finally:
            status_label_text.set(f"Extracted {outputpdf}")
            root.update()
            stxt = status_label_text.get()
            print(stxt)
            writer = None

        row_index += 1
        success_count += 1

    if (success_count > 0):
        status_label_text.set(f"successfully extracted {success_count} contract files")
    else:
        update_extraction_failed()

### GUI Setup ###

# Create the Tkinter window
root = tk.Tk()

# Set window title
root.title("PDF Extractor")

grid_rowspan = 12
# Create a canvas
canvas = tk.Canvas(root, width=600, height=400)
canvas.grid(columnspan=5, rowspan=grid_rowspan)

# Instructions label
instructions = tk.Label(root, text="Select a .pdf file and an .xlsx file to split the .pdf into separate entries", font="Raleway 14")
instructions.grid(columnspan=5, column=0, row=0)

# Font size for labels
font_size = 12

### Labels ###

# Labels for user inputs
entry_label = tk.Label(root, text="Pages per file", font=f"Raleway {font_size}")

enter_pdf_label = tk.Label(root, text="Source PDF", font=f"Raleway {font_size}")

enter_excel_label = tk.Label(root, text="Source Excel", font=f"Raleway {font_size}")

# Left-align labels and add padding
entry_label.grid(column=0, row=1, sticky="w",  padx=(60, 5), pady=5)
enter_pdf_label.grid(column=0, row=2, sticky="w",  padx=(60, 5), pady=5)
enter_excel_label.grid(column=0, row=5, sticky="w", padx=(60, 5), pady=5)

separator1 = ttk.Separator(root, orient='horizontal')
separator1.grid(row=1, column=0, columnspan=5, pady=(35, 0), padx=(5, 0), sticky='ew')

separator2 = ttk.Separator(root, orient='horizontal')
separator2.grid(row=3, column=0, columnspan=5, pady=(35, 0), padx=(5, 0), sticky='ew')

separator3 = ttk.Separator(root, orient='horizontal')
separator3.grid(row=6, column=0, columnspan=5, pady=(35, 0), padx=(5, 0), sticky='ew')

path_wrap_length = 550
path_width = 100
# Labels to display selected files
source_pdf_text = tk.StringVar()
source_pdf_text.set(os.path.abspath(str_pdf))
source_pdf_label = tk.Label(root, textvariable=source_pdf_text, font=f"Raleway {10}", wraplength=path_wrap_length)
source_pdf_label.grid(column=0, row=3, sticky="w", padx=(60, 15), pady=5, columnspan=5)

source_excel_text = tk.StringVar()
source_excel_text.set(os.path.abspath(str_wb))
source_excel_label = tk.Label(root, textvariable=source_excel_text, font=f"Raleway {10}", wraplength=path_wrap_length)
source_excel_label.grid(column=0, row=6, sticky="w", padx=(60, 15), pady=5, columnspan=5)

# Label to display status
status_label_text = tk.StringVar()
status_label = tk.Label(root, textvariable=status_label_text, font=f"Raleway {font_size}", wraplength=500)
status_label.grid(row=9, column=0, columnspan=5)

### Entry ###

# Entry widget for pages per contract
text_var = tk.StringVar()
entry_page_per_contract = tk.Entry(root, textvariable=text_var, borderwidth=5, width=10, relief=tk.GROOVE)
text_var.set(str(contract_pages_count))  # Assuming contract_pages_count is defined elsewhere
entry_page_per_contract.grid(column=1, row=1, columnspan=3, sticky="W")

### Checkbox ###

# Create a BooleanVar to associate with the Checkbutton
check_val = tk.BooleanVar()
# Checkbutton to overwrite existing PDF files
check_overwrite_pdf = tk.Checkbutton(root, variable=check_val, text="overwrite existing pdf files", font="Raleway")
check_overwrite_pdf.grid(column=0, row=7, columnspan=5)

### Buttons ###

# Browse buttons for selecting PDF and Excel files

default_text = "Browse to select"
pdf_browse_button = tk.Button(root, command=lambda: select_pdf_file(), text=default_text, font="Raleway", bg="#063b71", fg="white")
pdf_browse_button.grid(column=1, row=2,  columnspan=3, sticky="W")

excel_browse_button = tk.Button(root, command=lambda: select_excel_file(), text=default_text, font="Raleway", bg="#063b71", fg="white")
excel_browse_button.grid(column=1, row=5,  columnspan=3, sticky="W")

# Extract button
extract_button_text = "Extract Files"
extract_button = tk.Button(root, command=lambda: pdf_extract(), font="Arial", bg="#08AABD", fg="white", width=14, height=2, text=extract_button_text)
extract_button.grid(column=0, row=8, columnspan=5, pady=10)

# Start the Tkinter event loop
root.mainloop()
